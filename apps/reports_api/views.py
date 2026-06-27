from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import views, status
from apps.core.tenant_permissions import tenant_permission
from apps.core.permissions import IsSuperAdmin
from django.db.models import Count, Sum, Q
from django.utils import timezone
from datetime import datetime, timedelta
from apps.tenants_api.models import Tenant
from apps.billing_api.models import Invoice
from apps.auth_api.models import User
from apps.subscriptions_api.permissions import requires_feature
from .pagination import ReportsPagination
from apps.settings_api.policy_utils import (
    get_platform_commission_rate,
    calculate_platform_commission,
    calculate_platform_net_revenue,
)
from django.http import HttpResponse

def get_report_branch_id(request):
    branch_id = request.query_params.get('branch_id') or request.query_params.get('branch')
    user = request.user
    if user and getattr(user, 'is_authenticated', False) and not user.is_superuser:
        from apps.auth_api.role_utils import get_effective_role_api
        user_role = get_effective_role_api(user, tenant=getattr(request, 'tenant', user.tenant))
        if user_role != 'CLIENT_ADMIN' and hasattr(user, 'employee_profile') and user.employee_profile:
            if user.employee_profile.branch_id:
                branch_id = user.employee_profile.branch_id
    return branch_id

@api_view(['GET'])
@permission_classes([tenant_permission('reports_api.view_employee_reports')])
@requires_feature('reports')
def employee_report(request):
    """Reporte básico de empleados"""
    from apps.employees_api.models import Employee
    from apps.pos_api.models import Sale
    
    tenant = getattr(request, 'tenant', request.user.tenant)
    branch_id = get_report_branch_id(request)
    
    employee_filter = {'tenant': tenant}
    sale_filter = {'employee__tenant': tenant, 'employee__isnull': False}
    if branch_id:
        employee_filter['branch_id'] = branch_id
        sale_filter['branch_id'] = branch_id
        
    # Estadísticas básicas
    total_employees = Employee.objects.filter(**employee_filter).count()
    active_employees = Employee.objects.filter(**employee_filter, is_active=True).count()
    
    # Empleados por especialidad
    employees_by_specialty = Employee.objects.filter(
        **employee_filter, is_active=True
    ).values('specialty').annotate(count=Count('id'))
    
    # Top performers REALES desde ventas
    top_performers = Sale.objects.filter(
        **sale_filter
    ).values(
        'employee__user__full_name',
        'employee__specialty'
    ).annotate(
        total_sales=Sum('total'),
        sales_count=Count('id')
    ).order_by('-total_sales')[:5]
    
    top_performers_list = [{
        'employee_name': p['employee__user__full_name'],
        'sales': float(p['total_sales']),
        'sales_count': p['sales_count'],
        'specialty': p['employee__specialty'] or 'General'
    } for p in top_performers]
    
    return Response({
        'total_employees': total_employees,
        'active_employees': active_employees,
        'employees_by_specialty': list(employees_by_specialty),
        'top_performers': top_performers_list
    })

@api_view(['GET'])
@permission_classes([tenant_permission('reports_api.view_sales_reports')])
@requires_feature('reports')
def sales_report(request):
    """Reporte básico de ventas"""
    from apps.pos_api.models import Sale, SaleDetail
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    tenant = getattr(request, 'tenant', request.user.tenant)
    branch_id = get_report_branch_id(request)
    
    sale_filter = {'tenant': tenant}
    detail_filter = {'sale__tenant': tenant, 'content_type__model': 'service'}
    if branch_id:
        sale_filter['branch_id'] = branch_id
        detail_filter['sale__branch_id'] = branch_id

    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # Ventas totales
    total_sales = Sale.objects.filter(**sale_filter).aggregate(
        total=Sum('total')
    )['total'] or 0
    
    # Ventas del mes
    monthly_sales = Sale.objects.filter(
        **sale_filter,
        date_time__gte=month_start
    ).aggregate(total=Sum('total'))['total'] or 0
    
    # Ventas por día (últimos 7 días) — 1 query con TruncDate vs 7 queries individuales
    from django.db.models.functions import TruncDate
    week_ago = (now - timedelta(days=7)).replace(hour=0, minute=0, second=0, microsecond=0)
    daily_qs = Sale.objects.filter(
        **sale_filter,
        date_time__gte=week_ago
    ).annotate(
        day=TruncDate('date_time')
    ).values('day').annotate(
        total=Sum('total')
    ).order_by('day')
    
    daily_map = {d['day']: float(d['total']) for d in daily_qs}
    sales_by_day = []
    for i in range(6, -1, -1):
        day = (now - timedelta(days=i)).date()
        sales_by_day.append({
            'date': day.strftime('%Y-%m-%d'),
            'sales': daily_map.get(day, 0)
        })
    
    # Servicios más vendidos REALES
    top_services = SaleDetail.objects.filter(
        **detail_filter
    ).values('name').annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum('price')
    ).order_by('-total_revenue')[:5]
    
    return Response({
        'total_sales': float(total_sales),
        'monthly_sales': float(monthly_sales),
        'sales_by_day': sales_by_day,
        'top_services': list(top_services)
    })

@api_view(['GET'])
@permission_classes([tenant_permission('reports_api.view_kpi_dashboard')])
@requires_feature('reports')
def dashboard_stats(request):
    """Estadísticas para dashboard"""
    from django.core.cache import cache
    from apps.clients_api.models import Client
    from apps.employees_api.models import Employee
    from apps.pos_api.models import Sale
    from apps.appointments_api.models import Appointment
    
    tenant = getattr(request, 'tenant', request.user.tenant)
    branch_id = get_report_branch_id(request)
    
    # Cache key basado en tenant y branch
    cache_key = f'dashboard_stats_{getattr(tenant, "id", "none")}_branch_{branch_id or "all"}'
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return Response(cached_data)
    
    month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    employee_filter = {'tenant': tenant, 'is_active': True}
    appointment_filter = {'client__tenant': tenant, 'date_time__gte': month_start}
    sale_filter = {'tenant': tenant, 'date_time__gte': month_start}
    
    if branch_id:
        employee_filter['branch_id'] = branch_id
        appointment_filter['branch_id'] = branch_id
        sale_filter['branch_id'] = branch_id
        
    # Estadísticas básicas
    total_clients = Client.objects.filter(tenant=tenant, is_active=True).count()
    active_employees = Employee.objects.filter(**employee_filter).count()
    monthly_appointments = Appointment.objects.filter(**appointment_filter).count()
    monthly_revenue = Sale.objects.filter(**sale_filter).aggregate(total=Sum('total'))['total'] or 0

    data = {
        'total_clients': total_clients,
        'monthly_appointments': monthly_appointments,
        'monthly_revenue': float(monthly_revenue),
        'active_employees': active_employees
    }
    
    # Cache por 60 segundos
    cache.set(cache_key, data, 60)
    
    return Response(data)

@api_view(['GET'])
@permission_classes([tenant_permission('reports_api.view_sales_reports')])
@requires_feature('reports')
def reports_by_type(request):
    """Reportes por tipo (appointments, sales, etc.)"""
    report_type = request.GET.get('type', 'general')
    tenant = getattr(request, 'tenant', request.user.tenant)
    branch_id = get_report_branch_id(request)
    
    if report_type == 'appointments':
        from apps.appointments_api.models import Appointment
        from django.db.models import Count
        from datetime import datetime, timedelta
        
        # Últimos 6 meses de citas
        data = []
        labels = []
        for i in range(6):
            month_start = (timezone.now() - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            filters = {
                'client__tenant': tenant,
                'date_time__gte': month_start,
                'date_time__lte': month_end
            }
            if branch_id:
                filters['branch_id'] = branch_id
                
            count = Appointment.objects.filter(**filters).count()
            
            data.insert(0, count)
            labels.insert(0, month_start.strftime('%b'))
        
        return Response({'labels': labels, 'data': data})
        
    elif report_type == 'sales':
        from apps.pos_api.models import Sale
        from django.db.models import Sum
        from datetime import datetime, timedelta
        
        # Últimos 6 meses de ventas
        data = []
        labels = []
        for i in range(6):
            month_start = (timezone.now() - timedelta(days=30*i)).replace(day=1)
            month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            filters = {
                'tenant': tenant,
                'date_time__gte': month_start,
                'date_time__lte': month_end
            }
            if branch_id:
                filters['branch_id'] = branch_id
                
            total = Sale.objects.filter(**filters).aggregate(total=Sum('total'))['total'] or 0
            
            data.insert(0, float(total))
            labels.insert(0, month_start.strftime('%b'))
        
        return Response({'labels': labels, 'data': data})
    
    return Response({'data': []})

class AdminReportsView(views.APIView):
    """Reportes financieros para SuperAdmin"""
    permission_classes = [IsSuperAdmin]
    
    def get(self, request):
        try:
            period = request.GET.get('period', 'last_30_days')
            end_date = timezone.now()
            commission_rate = get_platform_commission_rate()

            # Calcular fechas según período (incluye rango custom real)
            if period == 'custom':
                start_raw = request.GET.get('start_date')
                end_raw = request.GET.get('end_date')
                if not start_raw or not end_raw:
                    return Response(
                        {'error': 'start_date y end_date son requeridos para período custom'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                try:
                    start_date = datetime.fromisoformat(start_raw).replace(
                        hour=0, minute=0, second=0, microsecond=0
                    )
                    end_date = datetime.fromisoformat(end_raw).replace(
                        hour=23, minute=59, second=59, microsecond=999999
                    )
                    if timezone.is_naive(start_date):
                        start_date = timezone.make_aware(start_date)
                    if timezone.is_naive(end_date):
                        end_date = timezone.make_aware(end_date)
                except ValueError:
                    return Response(
                        {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            elif period == 'last_7_days':
                start_date = end_date - timedelta(days=7)
            elif period == 'last_30_days':
                start_date = end_date - timedelta(days=30)
            elif period == 'last_3_months':
                start_date = end_date - timedelta(days=90)
            elif period == 'last_year':
                start_date = end_date - timedelta(days=365)
            else:
                start_date = end_date - timedelta(days=30)

            # Filtrar facturas del período
            invoices = Invoice.objects.filter(
                issued_at__gte=start_date,
                issued_at__lte=end_date
            )
            
            # Métricas principales
            total_revenue = invoices.filter(is_paid=True).aggregate(Sum('amount'))['amount__sum'] or 0
            platform_commission_amount = calculate_platform_commission(total_revenue)
            net_revenue = calculate_platform_net_revenue(total_revenue)
            pending_payments = invoices.filter(is_paid=False).aggregate(Sum('amount'))['amount__sum'] or 0
            overdue_invoices = invoices.filter(
                is_paid=False,
                due_date__lt=timezone.now()
            ).count()
            
            # Tendencia de ingresos REAL según el período seleccionado
            revenue_trend = []
            total_days = max(1, (end_date.date() - start_date.date()).days + 1)
            if total_days <= 45:
                cursor = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
                while cursor <= end_date:
                    day_end = cursor + timedelta(days=1)
                    day_revenue = Invoice.objects.filter(
                        issued_at__gte=cursor,
                        issued_at__lt=day_end,
                        is_paid=True
                    ).aggregate(Sum('amount'))['amount__sum'] or 0
                    day_commission = calculate_platform_commission(day_revenue)
                    day_net_revenue = calculate_platform_net_revenue(day_revenue)
                    revenue_trend.append({
                        'label': cursor.strftime('%d %b'),
                        'revenue': float(day_revenue),
                        'commission_amount': float(day_commission),
                        'net_revenue': float(day_net_revenue)
                    })
                    cursor = day_end
            else:
                cursor = start_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                while cursor <= end_date:
                    next_month = (cursor + timedelta(days=32)).replace(day=1)
                    month_revenue = Invoice.objects.filter(
                        issued_at__gte=cursor,
                        issued_at__lt=next_month,
                        issued_at__lte=end_date,
                        is_paid=True
                    ).aggregate(Sum('amount'))['amount__sum'] or 0
                    month_commission = calculate_platform_commission(month_revenue)
                    month_net_revenue = calculate_platform_net_revenue(month_revenue)
                    revenue_trend.append({
                        'label': cursor.strftime('%b %Y'),
                        'revenue': float(month_revenue),
                        'commission_amount': float(month_commission),
                        'net_revenue': float(month_net_revenue)
                    })
                    cursor = next_month
            
            # Top tenants por revenue - OPTIMIZADO: Una sola query
            top_tenants_data = Invoice.objects.filter(
                issued_at__gte=start_date,
                issued_at__lte=end_date,
                is_paid=True
            ).values(
                'tenant__name',
                'tenant__subscription_plan__name'
            ).annotate(
                total_revenue=Sum('amount')
            ).order_by('-total_revenue')[:5]
            
            top_tenants = [{
                'tenant_name': item['tenant__name'],
                'revenue': float(item['total_revenue']),
                'commission_amount': float(calculate_platform_commission(item['total_revenue'])),
                'net_revenue': float(calculate_platform_net_revenue(item['total_revenue'])),
                'plan': item['tenant__subscription_plan__name'] or 'Trial'
            } for item in top_tenants_data]
            
            return Response({
                'period': period,
                'total_revenue': float(total_revenue),
                'commission_rate': float(commission_rate),
                'platform_commission_amount': float(platform_commission_amount),
                'net_revenue': float(net_revenue),
                'pending_payments': float(pending_payments),
                'overdue_invoices': overdue_invoices,
                'revenue_trend': revenue_trend,
                'top_tenants': top_tenants[:5],
                'summary': {
                    'total_invoices': invoices.count(),
                    'paid_invoices': invoices.filter(is_paid=True).count(),
                    'average_invoice': float(total_revenue / invoices.count()) if invoices.count() > 0 else 0
                }
            })
            
        except Exception as e:
            return Response({
                'error': str(e),
                'message': 'Error al generar reportes'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([tenant_permission('reports_api.view_employee_reports')])
@requires_feature('reports')
def appointments_calendar_data(request):
    """Datos para calendario de citas con paginación"""
    from apps.appointments_api.models import Appointment
    
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    if not start_date or not end_date:
        return Response({'error': 'start y end son requeridos'}, status=400)
    
    try:
        start = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
    except ValueError:
        return Response({'error': 'Formato de fecha inválido'}, status=400)
    
    branch_id = get_report_branch_id(request)
    filters = {
        'client__tenant': getattr(request, 'tenant', request.user.tenant),
        'date_time__gte': start,
        'date_time__lte': end
    }
    if branch_id:
        filters['branch_id'] = branch_id

    appointments = Appointment.objects.filter(**filters).select_related('client', 'stylist', 'service')
    
    # Aplicar paginación si hay muchos resultados
    if appointments.count() > 100:
        paginator = ReportsPagination()
        page = paginator.paginate_queryset(appointments, request)
        if page is not None:
            appointments = page
    
    events = []
    for apt in appointments:
        color = {
            'scheduled': '#3498db',
            'completed': '#2ecc71', 
            'cancelled': '#e74c3c'
        }.get(apt.status, '#95a5a6')
        
        events.append({
            'id': apt.id,
            'title': f"{apt.client.full_name} - {apt.service.name if apt.service else 'Sin servicio'}",
            'start': apt.date_time.isoformat(),
            'end': (apt.date_time + timedelta(minutes=apt.service.duration if apt.service else 30)).isoformat(),
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'client': apt.client.full_name,
                'stylist': apt.stylist.full_name if apt.stylist else 'Sin asignar',
                'service': apt.service.name if apt.service else 'Sin servicio',
                'status': apt.status,
                'phone': apt.client.phone or 'N/A'
            }
        })
    
    return Response(events)

@api_view(['GET'])
@permission_classes([tenant_permission('reports_api.view_kpi_dashboard')])
@requires_feature('reports')
def kpi_dashboard(request):
    """KPIs principales para dashboard"""
    from django.core.cache import cache
    from apps.clients_api.models import Client
    from apps.employees_api.models import Employee
    from apps.pos_api.models import Sale
    from apps.appointments_api.models import Appointment
    from django.db.models import Sum, Count, Avg
    
    tenant = getattr(request, 'tenant', request.user.tenant)
    branch_id = get_report_branch_id(request)
    
    # Cache key basado en tenant y sucursal
    cache_key = f'kpi_dashboard_{getattr(tenant, "id", "none")}_branch_{branch_id or "all"}'
    cached_data = cache.get(cache_key)
    
    if cached_data:
        return Response(cached_data)
    
    today = timezone.now().date()
    month_start = today.replace(day=1)
    week_start = today - timedelta(days=today.weekday())
    
    sale_monthly_filter = {'tenant': tenant, 'date_time__date__gte': month_start}
    appointment_monthly_filter = {'client__tenant': tenant, 'date_time__date__gte': month_start}
    
    sale_weekly_filter = {'tenant': tenant, 'date_time__date__gte': week_start}
    appointment_weekly_filter = {'client__tenant': tenant, 'date_time__date__gte': week_start}
    
    employee_filter = {'tenant': tenant, 'is_active': True}
    
    if branch_id:
        sale_monthly_filter['branch_id'] = branch_id
        appointment_monthly_filter['branch_id'] = branch_id
        sale_weekly_filter['branch_id'] = branch_id
        appointment_weekly_filter['branch_id'] = branch_id
        employee_filter['branch_id'] = branch_id
    
    # KPIs del mes
    monthly_revenue = Sale.objects.filter(**sale_monthly_filter).aggregate(total=Sum('total'))['total'] or 0
    monthly_appointments = Appointment.objects.filter(**appointment_monthly_filter).count()
    
    # KPIs de la semana
    weekly_revenue = Sale.objects.filter(**sale_weekly_filter).aggregate(total=Sum('total'))['total'] or 0
    weekly_appointments = Appointment.objects.filter(**appointment_weekly_filter).count()
    
    # KPIs generales (Clients are global, Employees scoped)
    total_clients = Client.objects.filter(tenant=tenant, is_active=True).count()
    active_employees = Employee.objects.filter(**employee_filter).count()
    
    # Ticket promedio
    avg_ticket = Sale.objects.filter(**sale_monthly_filter).aggregate(avg=Avg('total'))['avg'] or 0
    
    data = {
        'monthly': {
            'revenue': float(monthly_revenue),
            'appointments': monthly_appointments,
            'avg_ticket': float(avg_ticket)
        },
        'weekly': {
            'revenue': float(weekly_revenue),
            'appointments': weekly_appointments
        },
        'totals': {
            'clients': total_clients,
            'employees': active_employees
        }
    }
    
    # Cache por 60 segundos
    cache.set(cache_key, data, 60)
    
    return Response(data)

@api_view(['GET'])
@permission_classes([tenant_permission('reports_api.view_kpi_dashboard')])
@requires_feature('reports')
def services_performance(request):
    """Rendimiento de servicios con paginación"""
    from apps.pos_api.models import SaleDetail
    from apps.services_api.models import Service
    from django.db.models import Sum, Count
    
    tenant = getattr(request, 'tenant', request.user.tenant)
    branch_id = get_report_branch_id(request)
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)
    
    filters = {
        'sale__tenant': tenant,
        'sale__date_time__gte': start_date,
        'content_type__model': 'service'
    }
    if branch_id:
        filters['sale__branch_id'] = branch_id

    # Top servicios por ventas
    services_data = SaleDetail.objects.filter(**filters).values('name').annotate(
        total_sales=Sum('price'),
        quantity_sold=Sum('quantity')
    ).order_by('-total_sales')
    
    # Aplicar paginación
    paginator = ReportsPagination()
    page = paginator.paginate_queryset(services_data, request)
    
    if page is not None:
        return paginator.get_paginated_response({
            'period_days': days,
            'services': list(page)
        })
    
    return Response({
        'period_days': days,
        'services': list(services_data)
    })

@api_view(['GET'])
@permission_classes([tenant_permission('reports_api.view_kpi_dashboard')])
@requires_feature('reports')
def client_analytics(request):
    """Análisis de clientes"""
    from apps.clients_api.models import Client
    from apps.appointments_api.models import Appointment
    from django.db.models import Count, Q
    
    tenant = getattr(request, 'tenant', request.user.tenant)
    branch_id = get_report_branch_id(request)
    month_start = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    current_month = timezone.now().month

    client_filter = {'tenant': tenant, 'is_active': True}
    new_client_filter = {'tenant': tenant, 'created_at__gte': month_start}
    birthday_filter = {'tenant': tenant, 'birthday__month': current_month, 'is_active': True}
    
    if branch_id:
        client_filter['client_appointments__branch_id'] = branch_id
        new_client_filter['client_appointments__branch_id'] = branch_id
        birthday_filter['client_appointments__branch_id'] = branch_id
    
    # Clientes por género
    gender_stats = Client.objects.filter(**client_filter).distinct().values('gender').annotate(
        count=Count('id', distinct=True)
    )
    
    # Clientes nuevos vs recurrentes (este mes)
    new_clients = Client.objects.filter(**new_client_filter).distinct().count()
    
    # Clientes con cumpleaños este mes
    birthday_clients = Client.objects.filter(**birthday_filter).distinct().count()
    
    return Response({
        'gender_distribution': list(gender_stats),
        'new_clients_this_month': new_clients,
        'birthday_clients': birthday_clients,
        'total_active': Client.objects.filter(**client_filter).distinct().count()
    })


@api_view(['GET'])
@permission_classes([tenant_permission('reports_api.view_sales_reports')])
@requires_feature('export_reports')
def export_report(request):
    """Exportar reportes a Excel - Solo Enterprise"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    report_type = request.GET.get('type', 'sales')
    tenant = getattr(request, 'tenant', request.user.tenant)
    branch_id = get_report_branch_id(request)

    wb = Workbook()
    ws = wb.active

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_header(ws, headers):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        ws.freeze_panes = 'A2'

    filters = {}
    if branch_id:
        filters['branch_id'] = branch_id

    if report_type == 'sales':
        from apps.pos_api.models import Sale
        headers = ['ID', 'Fecha', 'Cliente', 'Empleado', 'Total', 'Método de pago', 'Estado', 'Sucursal']
        style_header(ws, headers)

        base_filter = {'tenant': tenant}
        base_filter.update(filters)
        sales = Sale.objects.filter(**base_filter).select_related(
            'client', 'employee__user', 'branch'
        ).order_by('-date_time')[:5000]

        for row_idx, sale in enumerate(sales, 2):
            ws.cell(row=row_idx, column=1, value=sale.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=sale.date_time.strftime('%Y-%m-%d %H:%M') if sale.date_time else '').border = thin_border
            ws.cell(row=row_idx, column=3, value=sale.client.full_name if sale.client else '—').border = thin_border
            ws.cell(row=row_idx, column=4, value=sale.employee.user.full_name if sale.employee and sale.employee.user else '—').border = thin_border
            ws.cell(row=row_idx, column=5, value=float(sale.total)).border = thin_border
            ws.cell(row=row_idx, column=6, value=sale.payment_method or '—').border = thin_border
            ws.cell(row=row_idx, column=7, value=sale.status or '—').border = thin_border
            ws.cell(row=row_idx, column=8, value=sale.branch.name if sale.branch else '—').border = thin_border
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25

    elif report_type == 'appointments':
        from apps.appointments_api.models import Appointment
        headers = ['ID', 'Fecha', 'Cliente', 'Servicio', 'Empleado', 'Estado', 'Sucursal']
        style_header(ws, headers)

        base_filter = {'client__tenant': tenant}
        base_filter.update(filters)
        appointments = Appointment.objects.filter(**base_filter).select_related(
            'client', 'service', 'stylist', 'branch'
        ).order_by('-date_time')[:5000]

        for row_idx, apt in enumerate(appointments, 2):
            ws.cell(row=row_idx, column=1, value=apt.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=apt.date_time.strftime('%Y-%m-%d %H:%M') if apt.date_time else '').border = thin_border
            ws.cell(row=row_idx, column=3, value=apt.client.full_name if apt.client else '—').border = thin_border
            ws.cell(row=row_idx, column=4, value=apt.service.name if apt.service else '—').border = thin_border
            ws.cell(row=row_idx, column=5, value=apt.stylist.full_name if apt.stylist else '—').border = thin_border
            ws.cell(row=row_idx, column=6, value=apt.status or '—').border = thin_border
            ws.cell(row=row_idx, column=7, value=apt.branch.name if apt.branch else '—').border = thin_border
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 22
        ws.column_dimensions['E'].width = 25

    elif report_type == 'employees':
        from apps.employees_api.models import Employee
        headers = ['ID', 'Nombre', 'Email', 'Teléfono', 'Rol', 'Especialidad', 'Estado', 'Sucursal']
        style_header(ws, headers)

        base_filter = {'tenant': tenant}
        base_filter.update(filters)
        employees = Employee.objects.filter(**base_filter).select_related('user', 'branch')

        for row_idx, emp in enumerate(employees, 2):
            ws.cell(row=row_idx, column=1, value=emp.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=emp.user.full_name if emp.user else '—').border = thin_border
            ws.cell(row=row_idx, column=3, value=emp.user.email if emp.user else '—').border = thin_border
            ws.cell(row=row_idx, column=4, value=emp.phone or '—').border = thin_border
            ws.cell(row=row_idx, column=5, value=emp.user.business_role if emp.user else '—').border = thin_border
            ws.cell(row=row_idx, column=6, value=emp.specialty or '—').border = thin_border
            ws.cell(row=row_idx, column=7, value='Activo' if emp.is_active else 'Inactivo').border = thin_border
            ws.cell(row=row_idx, column=8, value=emp.branch.name if emp.branch else '—').border = thin_border
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30

    elif report_type == 'clients':
        from apps.clients_api.models import Client
        headers = ['ID', 'Nombre', 'Email', 'Teléfono', 'Género', 'Cumpleaños', 'Notas', 'Estado', 'Sucursal']
        style_header(ws, headers)

        base_filter = {'tenant': tenant}
        if branch_id:
            base_filter['branch_id'] = branch_id
        clients = Client.objects.filter(**base_filter).select_related('branch').order_by('-created_at')[:5000]

        for row_idx, client in enumerate(clients, 2):
            ws.cell(row=row_idx, column=1, value=client.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=client.full_name).border = thin_border
            ws.cell(row=row_idx, column=3, value=client.email or '—').border = thin_border
            ws.cell(row=row_idx, column=4, value=client.phone or '—').border = thin_border
            ws.cell(row=row_idx, column=5, value=dict(client.GENDER_CHOICES).get(client.gender, '—') if client.gender else '—').border = thin_border
            ws.cell(row=row_idx, column=6, value=client.birthday.strftime('%Y-%m-%d') if client.birthday else '—').border = thin_border
            ws.cell(row=row_idx, column=7, value=client.notes or '—').border = thin_border
            ws.cell(row=row_idx, column=8, value='Activo' if client.is_active else 'Inactivo').border = thin_border
            ws.cell(row=row_idx, column=9, value=client.branch.name if client.branch else '—').border = thin_border
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['G'].width = 30

    elif report_type == 'services':
        from apps.pos_api.models import SaleDetail
        headers = ['Servicio', 'Cantidad vendida', 'Ingresos totales']
        style_header(ws, headers)

        detail_filters = {
            'sale__tenant': tenant,
            'content_type__model': 'service'
        }
        if branch_id:
            detail_filters['sale__branch_id'] = branch_id

        services = SaleDetail.objects.filter(**detail_filters).values('name').annotate(
            total_sold=Sum('quantity'),
            total_revenue=Sum('price')
        ).order_by('-total_revenue')[:5000]

        for row_idx, svc in enumerate(services, 2):
            ws.cell(row=row_idx, column=1, value=svc['name']).border = thin_border
            ws.cell(row=row_idx, column=2, value=svc['total_sold']).border = thin_border
            ws.cell(row=row_idx, column=3, value=float(svc['total_revenue'])).border = thin_border
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['C'].width = 18

    elif report_type == 'products':
        from apps.pos_api.models import SaleDetail
        headers = ['Producto', 'Cantidad vendida', 'Ingresos totales']
        style_header(ws, headers)

        detail_filters = {
            'sale__tenant': tenant,
            'content_type__model': 'product'
        }
        if branch_id:
            detail_filters['sale__branch_id'] = branch_id

        products = SaleDetail.objects.filter(**detail_filters).values('name').annotate(
            total_sold=Sum('quantity'),
            total_revenue=Sum('price')
        ).order_by('-total_revenue')[:5000]

        for row_idx, prod in enumerate(products, 2):
            ws.cell(row=row_idx, column=1, value=prod['name']).border = thin_border
            ws.cell(row=row_idx, column=2, value=prod['total_sold']).border = thin_border
            ws.cell(row=row_idx, column=3, value=float(prod['total_revenue'])).border = thin_border
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['C'].width = 18

    elif report_type == 'cash_registers':
        from apps.pos_api.models import CashRegister
        headers = ['ID', 'Apertura', 'Cierre', 'Usuario', 'Monto Inicial', 'Ventas Efectivo', 'Monto Esperado', 'Monto Declarado', 'Diferencia', 'Estado', 'Sucursal']
        style_header(ws, headers)

        base_filter = {'tenant': tenant}
        base_filter.update(filters)
        registers = CashRegister.objects.filter(**base_filter).select_related(
            'user', 'branch'
        ).order_by('-opened_at')[:5000]

        for row_idx, reg in enumerate(registers, 2):
            sales_amount = float(reg.sales_amount)
            initial_cash = float(reg.initial_cash)
            final_cash = float(reg.final_cash)
            expected = initial_cash + sales_amount
            difference = final_cash - expected if not reg.is_open else 0.0

            ws.cell(row=row_idx, column=1, value=reg.id).border = thin_border
            ws.cell(row=row_idx, column=2, value=reg.opened_at.strftime('%Y-%m-%d %H:%M') if reg.opened_at else '').border = thin_border
            ws.cell(row=row_idx, column=3, value=reg.closed_at.strftime('%Y-%m-%d %H:%M') if reg.closed_at else '—').border = thin_border
            ws.cell(row=row_idx, column=4, value=getattr(reg.user, 'full_name', getattr(reg.user, 'email', '—'))).border = thin_border
            ws.cell(row=row_idx, column=5, value=initial_cash).border = thin_border
            ws.cell(row=row_idx, column=6, value=sales_amount).border = thin_border
            ws.cell(row=row_idx, column=7, value=final_cash if not reg.is_open else 0.0).border = thin_border
            ws.cell(row=row_idx, column=8, value=difference if not reg.is_open else 0.0).border = thin_border
            ws.cell(row=row_idx, column=9, value='Abierta' if reg.is_open else 'Cerrada').border = thin_border
            ws.cell(row=row_idx, column=10, value=reg.branch.name if reg.branch else '—').border = thin_border
        
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 25

    else:
        return Response({'error': f'Tipo de reporte inválido: {report_type}'}, status=400)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response
